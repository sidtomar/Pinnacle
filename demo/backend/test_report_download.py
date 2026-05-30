#!/usr/bin/env python3
"""Test the /report/download endpoint."""
import requests
import os

url = "http://127.0.0.1:8000/report/download"
print(f"Requesting: {url}")

try:
    resp = requests.get(url)
    print(f"Status: {resp.status_code}")
    print(f"Content-Type: {resp.headers.get('content-type')}")
    print(f"Content-Disposition: {resp.headers.get('content-disposition')}")
    print(f"Content-Length: {len(resp.content)} bytes")

    if resp.status_code == 200:
        # Save the file
        output_path = os.path.join(os.path.dirname(__file__), "test_report_output.xlsx")
        with open(output_path, "wb") as f:
            f.write(resp.content)
        print(f"\nSaved to: {output_path}")

        # Verify the file
        import openpyxl
        wb = openpyxl.load_workbook(output_path, data_only=True)
        print(f"\nSheet names: {wb.sheetnames}")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  {sheet_name}: {ws.max_row} rows x {ws.max_column} cols")

        print("\n[OK] Report generated successfully!")
    else:
        print(f"\nError: {resp.text}")

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
