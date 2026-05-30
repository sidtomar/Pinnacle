#!/usr/bin/env python3
"""Read the Excel template to understand its structure."""
import openpyxl

wb = openpyxl.load_workbook(
    r'C:\Users\siddhartha.tomar_tpr\Downloads\PinnacleIQ_PubMed_Repository_2026-05-27 1.xlsx',
    data_only=True
)

# Get actual column headers from row 3 (row 1-2 are title/subtitle)
ws = wb['All Publications']
print('=== ALL PUBLICATIONS - Column headers (row 3) ===')
headers = []
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=3, column=col).value
    headers.append(str(val) if val else f'col_{col}')
    print(f'  Col {col}: {val}')

# Print first 3 data rows with actual values
print('\n=== First 3 data rows ===')
for row in range(4, min(7, ws.max_row + 1)):
    print(f'\n  Row {row}:')
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=row, column=col).value
        if val is not None:
            print(f'    {headers[col-1]}: {str(val)[:100]}')

# Check Summary sheet structure
ws2 = wb['Summary']
print('\n\n=== SUMMARY SHEET ===')
for row in range(1, ws2.max_row + 1):
    row_vals = []
    for col in range(1, ws2.max_column + 1):
        val = ws2.cell(row=row, column=col).value
        if val is not None:
            row_vals.append(f'[{col}]={str(val)[:60]}')
    if row_vals:
        sep = '  |  '
        print(f'  Row {row}: {sep.join(row_vals)}')

# Check cell formatting/styles in All Publications
print('\n\n=== ALL PUBLICATIONS - Cell styles (row 1-3) ===')
ws3 = wb['All Publications']
for row in range(1, 4):
    for col in range(1, min(6, ws3.max_column + 1)):
        cell = ws3.cell(row=row, column=col)
        if cell.value:
            font = cell.font
            fill = cell.fill
            align = cell.alignment
            print(f'  [{row},{col}] val={str(cell.value)[:40]}  font={font.name}/{font.size}/{font.bold}/{font.color}  fill={fill.start_color}  align={align.horizontal}')

# Check Cardiology sheet for specialty-specific format
ws4 = wb['Cardiology']
print('\n\n=== CARDIOLOGY - Headers (row 3) ===')
for col in range(1, ws4.max_column + 1):
    val = ws4.cell(row=3, column=col).value
    print(f'  Col {col}: {val}')
print('\n  First data row (row 4):')
for col in range(1, ws4.max_column + 1):
    val = ws4.cell(row=4, column=col).value
    if val is not None:
        print(f'    Col {col}: {str(val)[:100]}')
