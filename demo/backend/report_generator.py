"""
Business Report Generator
=========================
Generates a professionally formatted Excel report matching the
PinnacleIQ PubMed Repository template structure.

Sheets:
  1. All Publications  - Every content card in the system
  2. Summary           - Specialty breakdown with counts
  3-N. Per-Specialty   - One tab per unique specialty
"""

import io
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Style constants (matching the template) ──────────────────────────────────
NAVY = "0D1F3C"
NAVY2 = "162E56"
GOLD = "B8912A"
YELLOW = "FFFF00"
WHITE = "FFFFFF"
GRAY_SUBTITLE = "888888"
LIGHT_GRAY = "F6F8FC"
BORDER_COLOR = "C8D2E4"

FONT_HEADER = Font(name="Arial", size=13, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name="Arial", size=8.5, color=GRAY_SUBTITLE)
FONT_COL_HEADER = Font(name="Arial", size=10, bold=True)
FONT_DATA = Font(name="Arial", size=9)
FONT_LINK = Font(name="Arial", size=9, color="1155CC", underline="single")
FONT_SUMMARY_TITLE = Font(name="Arial", size=14, bold=True, color=NAVY)
FONT_SUMMARY_SUB = Font(name="Arial", size=9, color=GRAY_SUBTITLE)
FONT_SUMMARY_HDR = Font(name="Arial", size=10, bold=True, color=WHITE)
FONT_SUMMARY_DATA = Font(name="Arial", size=10)
FONT_SUMMARY_TOTAL = Font(name="Arial", size=10, bold=True)
FONT_FIELD_GUIDE = Font(name="Arial", size=8, italic=True, color=GRAY_SUBTITLE)

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW)
FILL_NAVY2 = PatternFill("solid", fgColor=NAVY2)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT_GRAY)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_WRAP = Alignment(vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

COLUMNS = [
    ("#",                        4),
    ("Date",                    12),
    ("Specialty",               16),
    ("Therapy Area",            18),
    ("Sub-category",            16),
    ("Tags",                    22),
    ("Relevant Doctor Specialties", 28),
    ("WhatsApp Summary",        40),
    ("Title",                   45),
    ("PMID",                    12),
    ("DOI",                     25),
    ("Abstract",                50),
    ("Authors",                 30),
    ("PubMed Link",             32),
    ("Full Text Link (DOI)",    32),
]


def _na(val) -> str:
    """Return 'NA' for None/empty values, otherwise str(val)."""
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return "NA"
    return str(val)


def _map_item_to_row(item: dict, idx: int) -> list:
    """Convert a content_items record to a 15-column data row."""
    tags_raw = item.get("tags") or []
    if isinstance(tags_raw, list):
        tags = " | ".join(tags_raw) if tags_raw else "NA"
    else:
        tags = str(tags_raw) if tags_raw else "NA"

    date_val = item.get("publication_date") or (item.get("created_at", "") or "")[:10]

    return [
        idx,
        _na(date_val),
        _na(item.get("specialty")),
        _na(item.get("therapy_area")),
        _na(item.get("sub_category")),
        tags,
        _na(item.get("relevant_doctor_specialties")),
        _na(item.get("whatsapp_summary")),
        _na(item.get("title")),
        _na(item.get("pmid")),
        _na(item.get("doi")),
        _na(item.get("summary")),
        _na(item.get("authors")),
        _na(item.get("pubmed_link")),
        _na(item.get("full_text_link")),
    ]


def _write_title_rows(ws, title_text: str, subtitle_text: str, total_cols: int = 15):
    """Write the two-row title header (navy background, merged)."""
    # Row 1: Main title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    cell = ws.cell(row=1, column=1, value=title_text)
    cell.font = FONT_HEADER
    cell.fill = FILL_NAVY
    cell.alignment = ALIGN_CENTER
    for col in range(2, total_cols + 1):
        ws.cell(row=1, column=col).fill = FILL_NAVY

    # Row 2: Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    cell2 = ws.cell(row=2, column=1, value=subtitle_text)
    cell2.font = FONT_SUBTITLE
    cell2.alignment = ALIGN_CENTER

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 18


def _write_column_headers(ws, row_num: int = 3):
    """Write the yellow column-header row with autofilter."""
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=col_name)
        cell.font = FONT_COL_HEADER
        cell.fill = FILL_YELLOW
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[row_num].height = 22
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A{row_num}:{last_col}{row_num}"


def _write_data_rows(ws, rows: list, start_row: int = 4):
    """Write data rows with alternating shading and hyperlinks."""
    for r_idx, row_data in enumerate(rows):
        excel_row = start_row + r_idx
        for c_idx, val in enumerate(row_data):
            cell = ws.cell(row=excel_row, column=c_idx + 1, value=val)
            cell.font = FONT_DATA
            cell.alignment = ALIGN_WRAP
            cell.border = THIN_BORDER

            # Alternating row shading
            if r_idx % 2 == 1:
                cell.fill = FILL_LIGHT

            # Make links clickable
            col_name = COLUMNS[c_idx][0]
            if col_name in ("PubMed Link", "Full Text Link (DOI)") and val:
                cell.font = FONT_LINK
                cell.hyperlink = str(val)

        ws.row_dimensions[excel_row].height = 28


def _build_all_publications(wb: Workbook, items: list, report_date: str):
    """Build the 'All Publications' sheet."""
    ws = wb.active
    ws.title = "All Publications"

    specialties = sorted(set(i.get("specialty", "General") for i in items))
    spec_count = len(specialties)
    total = len(items)

    title = f"PinnacleIQ — All Medical Publications — {report_date}"
    subtitle = f"PubMed | {spec_count} Specialties | {total} articles | Tags & Doctor Relevance included"

    _write_title_rows(ws, title, subtitle)
    _write_column_headers(ws)

    rows = [_map_item_to_row(item, idx + 1) for idx, item in enumerate(items)]
    _write_data_rows(ws, rows)

    # Freeze panes below header
    ws.freeze_panes = "A4"


def _build_summary(wb: Workbook, items: list, report_date: str):
    """Build the 'Summary' sheet."""
    ws = wb.create_sheet("Summary")

    # Title (row 2, col B)
    ws.merge_cells("B2:I2")
    cell_title = ws.cell(row=2, column=2, value="PinnacleIQ — Publication Repository Summary")
    cell_title.font = FONT_SUMMARY_TITLE

    # Subtitle (row 3, col B)
    specialties = sorted(set(i.get("specialty", "General") for i in items))
    total = len(items)
    ws.merge_cells("B3:I3")
    cell_sub = ws.cell(row=3, column=2,
        value=f"{report_date}  |  Source: PubMed  |  Total: {total} articles  |  Classified by Specialty")
    cell_sub.font = FONT_SUMMARY_SUB

    # Table headers (row 5)
    summary_headers = ["Specialty", "Articles", "Sample Therapy Areas", "Sub-categories Found", "Top Relevant Specialties"]
    for col_idx, hdr in enumerate(summary_headers, 2):
        cell = ws.cell(row=5, column=col_idx, value=hdr)
        cell.font = FONT_SUMMARY_HDR
        cell.fill = FILL_NAVY2
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 50

    # Data rows per specialty
    row_num = 6
    for spec in specialties:
        spec_items = [i for i in items if i.get("specialty") == spec]
        count = len(spec_items)

        therapy_areas = sorted(set(i.get("therapy_area", "") for i in spec_items if i.get("therapy_area")))
        sub_cats = sorted(set(i.get("sub_category", "") for i in spec_items if i.get("sub_category")))
        doc_specs_all = []
        for i in spec_items:
            ds = i.get("relevant_doctor_specialties", "")
            if ds:
                doc_specs_all.extend([s.strip() for s in ds.replace("|", ",").split(",") if s.strip()])
        doc_specs = sorted(set(doc_specs_all))[:4]

        row_data = [
            spec,
            count,
            " • ".join(therapy_areas[:4]),
            " • ".join(sub_cats[:4]),
            " • ".join(doc_specs),
        ]

        for col_idx, val in enumerate(row_data, 2):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = FONT_SUMMARY_DATA
            cell.alignment = ALIGN_WRAP
            cell.border = THIN_BORDER
            if (row_num - 6) % 2 == 1:
                cell.fill = FILL_LIGHT

        ws.row_dimensions[row_num].height = 24
        row_num += 1

    # Total row
    ws.cell(row=row_num, column=2, value="TOTAL").font = FONT_SUMMARY_TOTAL
    ws.cell(row=row_num, column=2).border = THIN_BORDER
    ws.cell(row=row_num, column=3, value=total).font = FONT_SUMMARY_TOTAL
    ws.cell(row=row_num, column=3).border = THIN_BORDER

    # Field guide (2 rows below total)
    guide_row = row_num + 2
    ws.merge_cells(f"B{guide_row}:I{guide_row}")
    guide_text = (
        "FIELD GUIDE — Specialty: Primary clinical specialty (PinnacleIQ classification)  |  "
        "Therapy Area: Disease/treatment focus  |  Sub-category: Article type (RCT, Meta-Analysis, Review, etc.)  |  "
        "Tags: Searchable keywords  |  Relevant Doctor Specialties: Target physician audience"
    )
    ws.cell(row=guide_row, column=2, value=guide_text).font = FONT_FIELD_GUIDE


def _build_specialty_sheet(wb: Workbook, specialty: str, items: list, report_date: str):
    """Build a per-specialty sheet."""
    spec_items = [i for i in items if i.get("specialty") == specialty]
    if not spec_items:
        return

    # Sheet name max 31 chars, no special chars
    safe_name = specialty[:31].replace("/", "-").replace("&", "+")
    ws = wb.create_sheet(safe_name)

    title = f"PinnacleIQ — {specialty} Publications"
    subtitle = f"PubMed | {report_date} | {len(spec_items)} articles"

    _write_title_rows(ws, title, subtitle)
    _write_column_headers(ws)

    rows = [_map_item_to_row(item, idx + 1) for idx, item in enumerate(spec_items)]
    _write_data_rows(ws, rows)

    ws.freeze_panes = "A4"


def generate_business_report(items: list) -> io.BytesIO:
    """
    Generate the full PinnacleIQ Business Report as an in-memory Excel file.

    Args:
        items: List of content_items dicts from the store.

    Returns:
        BytesIO buffer containing the .xlsx file.
    """
    now = datetime.now(timezone.utc)
    report_date = now.strftime("%d %b %Y")

    wb = Workbook()

    # 1. All Publications
    _build_all_publications(wb, items, report_date)

    # 2. Summary
    _build_summary(wb, items, report_date)

    # 3. Per-specialty tabs
    specialties = sorted(set(i.get("specialty", "General") for i in items))
    for spec in specialties:
        _build_specialty_sheet(wb, spec, items, report_date)

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
